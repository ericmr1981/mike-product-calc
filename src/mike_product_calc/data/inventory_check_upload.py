"""Inventory check (盘点) and delivery (到货) Excel parser and Supabase uploader."""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mike_product_calc.data.supabase_client import MpcSupabaseClient


CHECK_EXPECTED_HEADERS = [
    "品项编码",
    "品项名称",
    "品项规格",
    "品项类别",
    "库存单位",
    "初盘数量",
    "复盘数量",
    "系统库存",
    "盘点差异",
    "库存均价",
    "差异金额",
    "明细状态",
]

CHECK_REQUIRED_TEXT_FIELDS = ["品项编码", "品项名称", "库存单位"]

CHECK_NUMERIC_FIELDS = [
    "初盘数量", "复盘数量", "系统库存", "盘点差异", "库存均价", "差异金额",
]

CHECK_FIELD_MAP = {
    "品项编码": "item_code",
    "品项名称": "item_name",
    "品项规格": "spec",
    "品项类别": "category",
    "库存单位": "unit",
    "初盘数量": "first_check_qty",
    "复盘数量": "second_check_qty",
    "系统库存": "system_qty",
    "盘点差异": "diff_qty",
    "库存均价": "avg_price",
    "差异金额": "diff_amount",
}

# Match both "盘点单明细 YYYY年MM月DD日..." and "盘点品项导出YYYY年MM月DD日HH时MM分SS秒..."
CHECK_FILENAME_RE = re.compile(
    r"盘点.+?(?P<y>\d{4})年(?P<m>\d{2})月(?P<d>\d{2})日"
    r"(?:(?P<h>\d{2})时(?P<min>\d{2})分(?P<s>\d{2})秒)?"
)

DELIVERY_EXPECTED_HEADERS = [
    "品项编码",
    "品项名称",
    "品项规格",
    "品项类别",
    "库存单位",
    "数量",
]

DELIVERY_FIELD_MAP = {
    "品项编码": "item_code",
    "品项名称": "item_name",
    "品项规格": "spec",
    "品项类别": "category",
    "库存单位": "unit",
    "数量": "delivery_qty",
}

DELIVERY_REQUIRED_TEXT_FIELDS = ["品项编码", "数量"]


class InventoryCheckUploadError(Exception):
    """Raised when check/delivery workbook structure cannot be processed."""


@dataclass
class PreparedRows:
    items: list[dict[str, Any]]
    warnings: dict[str, int]
    errors: dict[str, int]
    skipped_rows: int


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any, field_name: str, row_num: int) -> float:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise ValueError(f"row {row_num}: missing numeric field {field_name}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"row {row_num}: invalid numeric field {field_name}={value!r}") from exc


def parse_check_at_from_filename(filename: str) -> datetime | None:
    """Parse check datetime from inventory check filename.

    Returns None if the filename doesn't match expected patterns.
    """
    m = CHECK_FILENAME_RE.search(filename)
    if not m:
        return None
    year = int(m.group("y"))
    month = int(m.group("m"))
    day = int(m.group("d"))
    hour = int(m.group("h")) if m.group("h") else 0
    minute = int(m.group("min")) if m.group("min") else 0
    second = int(m.group("s")) if m.group("s") else 0
    return datetime(year, month, day, hour, minute, second, tzinfo=timezone.utc)


def parse_delivery_at_from_filename(filename: str) -> datetime | None:
    """Parse delivery datetime from filename like '到货记录20250529.xlsx'."""
    m = re.search(r"(\d{8})", filename)
    if not m:
        return None
    dt_str = m.group(1)
    return datetime(
        year=int(dt_str[:4]),
        month=int(dt_str[4:6]),
        day=int(dt_str[6:8]),
        tzinfo=timezone.utc,
    )


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _parse_sheet(file_path: Path, expected_headers: list[str]) -> tuple[list[str], list[list[Any]]]:
    """Load workbook, find sheet, extract headers and raw rows."""
    wb = load_workbook(file_path, data_only=True)
    sheet_name = wb.sheetnames[0]  # Use first sheet
    ws = wb[sheet_name]

    header_index: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        name = _normalize_text(raw)
        if name and name not in header_index:
            header_index[name] = col

    missing_headers = [h for h in expected_headers if h not in header_index]
    if missing_headers:
        raise InventoryCheckUploadError(
            f"header_mismatch: missing={missing_headers} in {file_path.name}"
        )

    rows_data: list[list[Any]] = []
    for row_num in range(2, ws.max_row + 1):
        row_cn = [ws.cell(row_num, header_index[h]).value for h in expected_headers]
        if all(_normalize_text(v) == "" for v in row_cn):
            continue
        rows_data.append(row_cn)
    return expected_headers, rows_data


def prepare_check_inventory_rows(file_path: Path) -> PreparedRows:
    """Parse inventory check xlsx and return validated rows."""
    headers, raw_rows = _parse_sheet(file_path, CHECK_EXPECTED_HEADERS)

    warnings: dict[str, int] = {}
    errors: dict[str, int] = {}
    items: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    skipped_rows = 0

    for row_num, row_cn in enumerate(raw_rows, start=2):
        row_dict = dict(zip(headers, row_cn))
        invalid = False

        for h in CHECK_REQUIRED_TEXT_FIELDS:
            if _normalize_text(row_dict.get(h, "")) == "":
                errors[f"missing_required_{CHECK_FIELD_MAP[h]}"] = (
                    errors.get(f"missing_required_{CHECK_FIELD_MAP[h]}", 0) + 1
                )
                invalid = True

        numeric_values: dict[str, float] = {}
        if not invalid:
            for h in CHECK_NUMERIC_FIELDS:
                try:
                    numeric_values[h] = _to_float(row_dict.get(h, 0), h, row_num)
                except ValueError:
                    errors[f"invalid_numeric_{CHECK_FIELD_MAP[h]}"] = (
                        errors.get(f"invalid_numeric_{CHECK_FIELD_MAP[h]}", 0) + 1
                    )
                    invalid = True
                    break

        item_code = _normalize_text(row_dict.get("品项编码", ""))
        if not invalid and item_code in seen_codes:
            errors["duplicate_item_code_in_batch"] = errors.get("duplicate_item_code_in_batch", 0) + 1
            invalid = True

        if invalid:
            skipped_rows += 1
            continue
        seen_codes.add(item_code)

        data_warnings: list[str] = []
        if numeric_values.get("系统库存", 0) < 0:
            data_warnings.append("negative_stock")
            warnings["negative_stock"] = warnings.get("negative_stock", 0) + 1

        item = {
            "item_code": item_code,
            "item_name": _normalize_text(row_dict.get("品项名称", "")),
            "spec": _normalize_text(row_dict.get("品项规格", "")) or None,
            "unit": _normalize_text(row_dict.get("库存单位", "")),
            "category": _normalize_text(row_dict.get("品项类别", "")) or None,
            "system_qty": numeric_values["系统库存"],
            "first_check_qty": numeric_values.get("初盘数量"),
            "second_check_qty": numeric_values.get("复盘数量"),
            "diff_qty": numeric_values.get("盘点差异"),
            "avg_price": numeric_values.get("库存均价"),
            "diff_amount": numeric_values.get("差异金额"),
            "data_warnings": data_warnings,
        }
        items.append(item)

    return PreparedRows(items=items, warnings=warnings, errors=errors, skipped_rows=skipped_rows)


def prepare_delivery_rows(file_path: Path) -> PreparedRows:
    """Parse delivery xlsx and return validated rows."""
    headers, raw_rows = _parse_sheet(file_path, DELIVERY_EXPECTED_HEADERS)

    warnings: dict[str, int] = {}
    errors: dict[str, int] = {}
    items: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    skipped_rows = 0

    for row_num, row_cn in enumerate(raw_rows, start=2):
        row_dict = dict(zip(headers, row_cn))
        invalid = False

        item_code = _normalize_text(row_dict.get("品项编码", ""))
        if not item_code:
            errors["missing_required_item_code"] = errors.get("missing_required_item_code", 0) + 1
            invalid = True

        if not invalid:
            try:
                delivery_qty = _to_float(row_dict.get("数量", 0), "数量", row_num)
            except ValueError:
                errors["invalid_numeric_delivery_qty"] = errors.get("invalid_numeric_delivery_qty", 0) + 1
                invalid = True

        if not invalid and item_code in seen_codes:
            errors["duplicate_item_code_in_batch"] = errors.get("duplicate_item_code_in_batch", 0) + 1
            invalid = True

        if invalid:
            skipped_rows += 1
            continue
        seen_codes.add(item_code)

        item = {
            "item_code": item_code,
            "item_name": _normalize_text(row_dict.get("品项名称", "")) or None,
            "spec": _normalize_text(row_dict.get("品项规格", "")) or None,
            "unit": _normalize_text(row_dict.get("库存单位", "")) or None,
            "category": _normalize_text(row_dict.get("品项类别", "")) or None,
            "delivery_qty": delivery_qty,
        }
        items.append(item)

    return PreparedRows(items=items, warnings=warnings, errors=errors, skipped_rows=skipped_rows)


def _chunked(rows: list[dict[str, Any]], size: int = 500) -> list[list[dict[str, Any]]]:
    return [rows[i : i + size] for i in range(0, len(rows), size)]


def sync_check_inventory_file(
    client: MpcSupabaseClient | None,
    file_path: Path,
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Parse and upload an inventory check file to Supabase."""
    if not file_path.exists():
        raise InventoryCheckUploadError(f"file_not_found: {file_path}")

    prepared = prepare_check_inventory_rows(file_path)
    check_at = parse_check_at_from_filename(file_path.name)
    check_at_iso = check_at.isoformat() if check_at else datetime.now(timezone.utc).isoformat()

    if client is not None:
        duplicate = client.find_check_batch(source_filename=file_path.name)
        if duplicate:
            return {
                "file": str(file_path),
                "source_filename": file_path.name,
                "status": "skipped_duplicate",
                "batch_id": duplicate.get("id"),
                "inserted_rows": 0,
                "skipped_rows": 0,
                "warning_count": 0,
                "error_count": 0,
            }

    error_count = sum(prepared.errors.values())
    warning_count = sum(prepared.warnings.values())
    status = "imported"
    if not prepared.items:
        status = "failed"
    elif error_count > 0:
        status = "partial"

    if dry_run:
        return {
            "file": str(file_path),
            "source_filename": file_path.name,
            "status": "dry_run",
            "inserted_rows": len(prepared.items),
            "skipped_rows": prepared.skipped_rows,
            "warning_count": warning_count,
            "error_count": error_count,
            "warning_summary": prepared.warnings,
            "error_summary": prepared.errors,
        }

    if client is None:
        raise InventoryCheckUploadError("client_required_for_write")

    batch = client.create_check_batch(
        {
            "source_filename": file_path.name,
            "check_at": check_at_iso,
            "row_count": 0,
            "status": "imported",
            "error_count": 0,
            "error_summary": [],
        }
    )
    batch_id = batch["id"]

    inserted_rows = 0
    for group in _chunked(prepared.items, size=500):
        payload = [{**row, "batch_id": batch_id} for row in group]
        client.insert_check_items(payload)
        inserted_rows += len(group)

    client.update_check_batch(
        batch_id,
        {
            "row_count": inserted_rows,
            "status": status,
            "error_count": error_count,
            "error_summary": prepared.errors,
        },
    )

    return {
        "file": str(file_path),
        "source_filename": file_path.name,
        "status": status,
        "batch_id": batch_id,
        "inserted_rows": inserted_rows,
        "skipped_rows": prepared.skipped_rows,
        "warning_count": warning_count,
        "error_count": error_count,
        "warning_summary": prepared.warnings,
        "error_summary": prepared.errors,
    }


def sync_delivery_file(
    client: MpcSupabaseClient | None,
    file_path: Path,
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Parse and upload a delivery file to Supabase."""
    if not file_path.exists():
        raise InventoryCheckUploadError(f"file_not_found: {file_path}")

    prepared = prepare_delivery_rows(file_path)
    delivery_at = parse_delivery_at_from_filename(file_path.name)
    delivery_at_iso = delivery_at.isoformat() if delivery_at else datetime.now(timezone.utc).isoformat()

    if client is not None:
        duplicate = client.find_delivery_batch(source_filename=file_path.name)
        if duplicate:
            return {
                "file": str(file_path),
                "source_filename": file_path.name,
                "status": "skipped_duplicate",
                "batch_id": duplicate.get("id"),
                "inserted_rows": 0,
                "skipped_rows": 0,
                "warning_count": 0,
                "error_count": 0,
            }

    error_count = sum(prepared.errors.values())
    warning_count = sum(prepared.warnings.values())
    status = "imported"
    if not prepared.items:
        status = "failed"
    elif error_count > 0:
        status = "partial"

    if dry_run:
        return {
            "file": str(file_path),
            "source_filename": file_path.name,
            "status": "dry_run",
            "inserted_rows": len(prepared.items),
            "skipped_rows": prepared.skipped_rows,
            "warning_count": warning_count,
            "error_count": error_count,
            "warning_summary": prepared.warnings,
            "error_summary": prepared.errors,
        }

    if client is None:
        raise InventoryCheckUploadError("client_required_for_write")

    batch = client.create_delivery_batch(
        {
            "source_filename": file_path.name,
            "delivery_at": delivery_at_iso,
            "row_count": 0,
            "status": "imported",
            "error_count": 0,
            "error_summary": [],
        }
    )
    batch_id = batch["id"]

    inserted_rows = 0
    for group in _chunked(prepared.items, size=500):
        payload = [{**row, "batch_id": batch_id} for row in group]
        client.insert_delivery_items(payload)
        inserted_rows += len(group)

    client.update_delivery_batch(
        batch_id,
        {
            "row_count": inserted_rows,
            "status": status,
            "error_count": error_count,
            "error_summary": prepared.errors,
        },
    )

    return {
        "file": str(file_path),
        "source_filename": file_path.name,
        "status": status,
        "batch_id": batch_id,
        "inserted_rows": inserted_rows,
        "skipped_rows": prepared.skipped_rows,
        "warning_count": warning_count,
        "error_count": error_count,
        "warning_summary": prepared.warnings,
        "error_summary": prepared.errors,
    }

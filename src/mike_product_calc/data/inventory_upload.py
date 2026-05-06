"""Inventory snapshot uploader for warehouse stock export xlsx files."""

from __future__ import annotations

import hashlib
import math
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from mike_product_calc.data.supabase_client import MpcSupabaseClient


EXPECTED_SHEET_NAME = "仓库库存导出"
EXPECTED_HEADERS = [
    "品项编码",
    "品项名称",
    "规格",
    "单位",
    "二级品项类别",
    "一级品项类别",
    "品项属性名",
    "仓库名称",
    "仓库编码",
    "库存量",
    "可用量",
    "占用量",
    "预计出库量",
    "预计入库量",
    "现存金额",
    "库存单价",
]
REQUIRED_TEXT_FIELDS = [
    "品项编码",
    "品项名称",
    "单位",
    "二级品项类别",
    "仓库名称",
    "仓库编码",
]
NUMERIC_FIELDS = [
    "库存量",
    "可用量",
    "占用量",
    "预计出库量",
    "预计入库量",
    "现存金额",
    "库存单价",
]
FIELD_MAP = {
    "品项编码": "item_code",
    "品项名称": "item_name",
    "规格": "spec",
    "单位": "unit",
    "二级品项类别": "category_lv2",
    "一级品项类别": "category_lv1",
    "品项属性名": "item_attribute_name",
    "仓库名称": "warehouse_name",
    "仓库编码": "warehouse_code",
    "库存量": "stock_qty",
    "可用量": "available_qty",
    "占用量": "occupied_qty",
    "预计出库量": "expected_out_qty",
    "预计入库量": "expected_in_qty",
    "现存金额": "current_amount",
    "库存单价": "stock_unit_price",
}
SNAPSHOT_FILENAME_RE = re.compile(
    r"仓库库存导出(?P<y>\d{4})年(?P<m>\d{2})月(?P<d>\d{2})日(?P<h>\d{2})时(?P<min>\d{2})分(?P<s>\d{2})秒\.xlsx$"
)


class InventoryUploadError(Exception):
    """Raised when workbook structure cannot be processed."""


@dataclass
class PreparedRows:
    items: list[dict[str, Any]]
    warnings: dict[str, int]
    errors: dict[str, int]
    skipped_rows: int


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any, field_name: str, row_num: int) -> float:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise ValueError(f"row {row_num}: missing numeric field {field_name}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"row {row_num}: invalid numeric field {field_name}={value!r}") from exc


def parse_snapshot_at(filename: str, tz_name: str = "Asia/Shanghai") -> datetime:
    """Parse snapshot datetime from exported xlsx filename."""
    m = SNAPSHOT_FILENAME_RE.match(filename)
    if m:
        dt_local = datetime(
            year=int(m.group("y")),
            month=int(m.group("m")),
            day=int(m.group("d")),
            hour=int(m.group("h")),
            minute=int(m.group("min")),
            second=int(m.group("s")),
            tzinfo=ZoneInfo(tz_name),
        )
        return dt_local.astimezone(timezone.utc)
    return datetime.now(timezone.utc)


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def prepare_inventory_rows(
    file_path: Path,
    *,
    sheet_name: str = EXPECTED_SHEET_NAME,
    amount_tolerance: float = 0.05,
) -> PreparedRows:
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise InventoryUploadError(f"sheet_not_found: {sheet_name}")
    ws = wb[sheet_name]

    header_index: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        name = _normalize_text(raw)
        if name and name not in header_index:
            header_index[name] = col

    missing_headers = [h for h in EXPECTED_HEADERS if h not in header_index]
    if missing_headers:
        raise InventoryUploadError(f"header_mismatch: missing={missing_headers}")

    warnings: dict[str, int] = {}
    errors: dict[str, int] = {}
    items: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, str]] = set()
    skipped_rows = 0

    def add_warning(code: str) -> None:
        warnings[code] = warnings.get(code, 0) + 1

    def add_error(code: str) -> None:
        errors[code] = errors.get(code, 0) + 1

    for row_num in range(2, ws.max_row + 1):
        row_cn: dict[str, Any] = {}
        for h in EXPECTED_HEADERS:
            row_cn[h] = ws.cell(row_num, header_index[h]).value

        if all(_normalize_text(v) == "" for v in row_cn.values()):
            continue

        invalid = False
        for h in REQUIRED_TEXT_FIELDS:
            if _normalize_text(row_cn[h]) == "":
                add_error(f"missing_required_{FIELD_MAP[h]}")
                invalid = True
        numeric_values: dict[str, float] = {}
        if not invalid:
            for h in NUMERIC_FIELDS:
                try:
                    numeric_values[h] = _to_float(row_cn[h], h, row_num)
                except ValueError:
                    add_error(f"invalid_numeric_{FIELD_MAP[h]}")
                    invalid = True
                    break

        item_code = _normalize_text(row_cn["品项编码"])
        warehouse_code = _normalize_text(row_cn["仓库编码"])
        key = (item_code, warehouse_code)
        if not invalid and key in seen_keys:
            add_error("duplicate_item_warehouse_in_batch")
            invalid = True

        if invalid:
            skipped_rows += 1
            continue
        seen_keys.add(key)

        data_warnings: list[str] = []
        if numeric_values["库存量"] < 0:
            data_warnings.append("negative_stock")
            add_warning("negative_stock")
        if _normalize_text(row_cn["规格"]) == "":
            data_warnings.append("empty_spec")
            add_warning("empty_spec")
        calc_amount = numeric_values["库存量"] * numeric_values["库存单价"]
        if not math.isclose(calc_amount, numeric_values["现存金额"], rel_tol=1e-6, abs_tol=amount_tolerance):
            data_warnings.append("amount_mismatch")
            add_warning("amount_mismatch")

        item = {
            "item_code": item_code,
            "item_name": _normalize_text(row_cn["品项名称"]),
            "spec": _normalize_text(row_cn["规格"]) or None,
            "unit": _normalize_text(row_cn["单位"]),
            "category_lv2": _normalize_text(row_cn["二级品项类别"]),
            "category_lv1": _normalize_text(row_cn["一级品项类别"]) or None,
            "item_attribute_name": _normalize_text(row_cn["品项属性名"]) or None,
            "warehouse_name": _normalize_text(row_cn["仓库名称"]),
            "warehouse_code": warehouse_code,
            "stock_qty": numeric_values["库存量"],
            "available_qty": numeric_values["可用量"],
            "occupied_qty": numeric_values["占用量"],
            "expected_out_qty": numeric_values["预计出库量"],
            "expected_in_qty": numeric_values["预计入库量"],
            "current_amount": numeric_values["现存金额"],
            "stock_unit_price": numeric_values["库存单价"],
            "is_negative_stock": "negative_stock" in data_warnings,
            "has_amount_mismatch": "amount_mismatch" in data_warnings,
            "data_warnings": data_warnings,
        }
        items.append(item)

    return PreparedRows(items=items, warnings=warnings, errors=errors, skipped_rows=skipped_rows)


def _chunked(rows: list[dict[str, Any]], size: int = 500) -> list[list[dict[str, Any]]]:
    return [rows[i : i + size] for i in range(0, len(rows), size)]


def sync_inventory_file(
    client: MpcSupabaseClient | None,
    file_path: Path,
    *,
    sheet_name: str = EXPECTED_SHEET_NAME,
    timezone_name: str = "Asia/Shanghai",
    dry_run: bool = False,
) -> dict[str, Any]:
    if not file_path.exists():
        raise InventoryUploadError(f"file_not_found: {file_path}")

    sha = file_sha256(file_path)
    snapshot_at = parse_snapshot_at(file_path.name, timezone_name).isoformat()
    prepared = prepare_inventory_rows(file_path, sheet_name=sheet_name)

    if client is not None:
        duplicate = client.find_inventory_batch(source_filename=file_path.name, source_file_sha256=sha)
        if duplicate:
            return {
                "file": str(file_path),
                "source_filename": file_path.name,
                "source_file_sha256": sha,
                "snapshot_at": snapshot_at,
                "status": "skipped_duplicate",
                "batch_id": duplicate.get("id"),
                "inserted_rows": 0,
                "skipped_rows": 0,
                "warning_count": 0,
                "error_count": 0,
            }

    error_count = sum(prepared.errors.values())
    warning_count = sum(prepared.warnings.values())
    status = "imported"
    if not prepared.items:
        status = "failed"
    elif error_count > 0:
        status = "partial"

    if dry_run:
        return {
            "file": str(file_path),
            "source_filename": file_path.name,
            "source_file_sha256": sha,
            "snapshot_at": snapshot_at,
            "status": "dry_run",
            "inserted_rows": len(prepared.items),
            "skipped_rows": prepared.skipped_rows,
            "warning_count": warning_count,
            "error_count": error_count,
            "warning_summary": prepared.warnings,
            "error_summary": prepared.errors,
        }

    if client is None:
        raise InventoryUploadError("client_required_for_write")

    batch = client.create_inventory_batch(
        {
            "source_filename": file_path.name,
            "source_sheet_name": sheet_name,
            "source_file_sha256": sha,
            "snapshot_at": snapshot_at,
            "row_count": 0,
            "status": "imported",
            "warning_count": 0,
            "error_count": 0,
            "warning_summary": [],
            "error_summary": [],
        }
    )
    batch_id = batch["id"]

    inserted_rows = 0
    for group in _chunked(prepared.items, size=500):
        payload = [{**row, "batch_id": batch_id} for row in group]
        client.insert_inventory_items(payload)
        inserted_rows += len(group)

    client.update_inventory_batch(
        batch_id,
        {
            "row_count": inserted_rows,
            "status": status,
            "warning_count": warning_count,
            "error_count": error_count,
            "warning_summary": prepared.warnings,
            "error_summary": prepared.errors,
        },
    )

    return {
        "file": str(file_path),
        "source_filename": file_path.name,
        "source_file_sha256": sha,
        "snapshot_at": snapshot_at,
        "status": status,
        "batch_id": batch_id,
        "inserted_rows": inserted_rows,
        "skipped_rows": prepared.skipped_rows,
        "warning_count": warning_count,
        "error_count": error_count,
        "warning_summary": prepared.warnings,
        "error_summary": prepared.errors,
    }


def discover_inventory_files(path: Path, pattern: str = "仓库库存导出*.xlsx") -> list[Path]:
    if path.is_file():
        return [path]
    if not path.exists():
        raise InventoryUploadError(f"path_not_found: {path}")
    files = sorted(path.glob(pattern))
    if not files:
        raise InventoryUploadError(f"no_files_matched: {pattern}")
    return files
